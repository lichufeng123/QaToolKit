from __future__ import annotations

from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from openpyxl import Workbook
from openpyxl.styles import Font

from qatoolkit.testcase_import.service import _module_id, import_testcases_from_excel


class TestcaseImportTests(unittest.TestCase):
    def test_module_path_prefers_workflow_child_module(self) -> None:
        self.assertEqual(_module_id("工作流"), 124)
        self.assertEqual(_module_id("工作流/局部重绘"), 175)
        self.assertEqual(_module_id("/工作流/局部重绘"), 175)
        self.assertEqual(_module_id("工作流/智能分镜脚本节点"), 174)
        self.assertEqual(_module_id("工作流/图像检测"), 178)
        self.assertEqual(_module_id("个人中心/分享链接"), 179)
        self.assertEqual(_module_id("收费系统"), 180)
        self.assertEqual(_module_id("收费系统/会员订阅系统"), 181)
        self.assertEqual(_module_id("收费系统/会员权益配置"), 182)

    def test_dry_run_maps_sheet_module_and_puts_single_expect_on_last_step(self) -> None:
        with TemporaryDirectory() as tmp:
            workbook_path = Path(tmp) / "cases.xlsx"
            report_path = Path(tmp) / "report.json"

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "密码登录"
            sheet.append(["用例标题", "步骤", "预期", "优先级", "类型"])
            sheet.append(["密码正确性校验", "输入账号\n输入密码\n点击登录", "登录成功", "3", "功能"])
            workbook.save(workbook_path)

            result = import_testcases_from_excel(
                workbook_path=workbook_path,
                base_url="",
                dry_run=True,
                output_path=report_path,
            )

            case = result.sheets[0].cases[0]
            self.assertEqual(result.parsed_count, 1)
            self.assertEqual(result.success_count, 1)
            self.assertEqual(result.sheets[0].module_name, "密码登录")
            self.assertEqual(result.sheets[0].module_id, 172)
            self.assertEqual(case.payload["module"], 172)
            self.assertEqual(case.payload["type"], "feature")
            self.assertEqual(case.payload["steps"], ["输入账号", "输入密码", "点击登录"])
            self.assertEqual(case.payload["expects"], ["", "", "登录成功"])
            self.assertTrue(report_path.exists())

    def test_dry_run_maps_module_column_to_workflow_child_module(self) -> None:
        with TemporaryDirectory() as tmp:
            workbook_path = Path(tmp) / "cases.xlsx"
            report_path = Path(tmp) / "report.json"

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "测试用例"
            sheet.append(["测试模块", "用例标题", "测试步骤", "预期结果"])
            sheet.append(["工作流/局部重绘", "验证局部重绘入口", "打开工作流\n点击局部重绘", "进入局部重绘节点"])
            workbook.save(workbook_path)

            result = import_testcases_from_excel(
                workbook_path=workbook_path,
                base_url="",
                dry_run=True,
                output_path=report_path,
            )

            case = result.sheets[0].cases[0]
            self.assertEqual(result.sheets[0].module_name, "工作流/局部重绘")
            self.assertEqual(result.sheets[0].module_id, 175)
            self.assertEqual(case.payload["module"], 175)

    def test_dry_run_skips_struck_through_rows(self) -> None:
        with TemporaryDirectory() as tmp:
            workbook_path = Path(tmp) / "cases.xlsx"
            report_path = Path(tmp) / "report.json"

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "登录"
            sheet.append(["用例标题", "步骤", "预期"])
            sheet.append(["废弃用例", "打开页面", "显示页面"])
            sheet["A2"].font = Font(strike=True)
            workbook.save(workbook_path)

            result = import_testcases_from_excel(
                workbook_path=workbook_path,
                base_url="",
                dry_run=True,
                output_path=report_path,
            )

            self.assertEqual(result.parsed_count, 0)
            self.assertEqual(result.skipped_count, 1)
            self.assertEqual(result.success_count, 0)


if __name__ == "__main__":
    unittest.main()
