from __future__ import annotations

from dataclasses import asdict, dataclass, field
from datetime import datetime
import json
from pathlib import Path
import re
import shutil
import tempfile
from typing import Any
from zipfile import ZipFile

from openpyxl import load_workbook
import requests

from ..shared.environment import load_dotenv_file
from ..shared.paths import project_root as find_project_root


DEFAULT_PRODUCT_ID = 8
MODULE_ID_MAP = {
    "首页": 121,
    "AI员工": 122,
    "AI群组": 123,
    "工作流": 124,
    "公共支持服务": 125,
    "个人中心": 126,
    "赛点详情": 158,
    "资产库": 159,
    "需求池": 168,
    "登录": 172,
}

MODULE_NAME_ALIASES = {
    "首页": {"首页"},
    "AI员工": {"AI员工", "员工", "智能体", "AI智能体"},
    "AI群组": {"AI群组", "群组"},
    "工作流": {"工作流"},
    "公共支持服务": {"公共支持服务", "公共服务", "支持服务"},
    "个人中心": {"个人中心", "我的", "用户中心"},
    "赛点详情": {"赛点详情"},
    "资产库": {"资产库", "素材库"},
    "需求池": {"需求池"},
    "登录": {"登录", "密码登录", "验证码登录", "手机号登录"},
}

HEADER_ALIASES = {
    "title": {"用例标题", "标题", "测试标题", "用例名称", "名称", "title"},
    "module": {"模块", "所属模块", "功能模块", "测试模块", "module"},
    "story": {"相关需求", "需求", "story"},
    "pri": {"优先级", "优先级别", "pri"},
    "type": {"用例类型", "类型", "type"},
    "precondition": {"前置条件", "前提条件", "precondition"},
    "steps": {"步骤", "测试步骤", "操作步骤", "执行步骤", "step", "steps"},
    "expects": {"预期", "预期结果", "期望结果", "expects", "expect"},
    "project": {"所属项目", "项目", "project"},
    "execution": {"所属执行", "执行", "execution"},
}

CASE_TYPE_MAP = {
    "unit": "unit",
    "单元测试": "unit",
    "interface": "interface",
    "接口测试": "interface",
    "feature": "feature",
    "功能测试": "feature",
    "功能": "feature",
    "install": "install",
    "安装部署": "install",
    "config": "config",
    "配置相关": "config",
    "performance": "performance",
    "性能测试": "performance",
    "security": "security",
    "安全相关": "security",
    "other": "other",
    "其他": "other",
}

STEP_SPLIT_PATTERN = re.compile(r"[\r\n]+")
LINE_PREFIX_PATTERN = re.compile(r"^\s*(?:\d+[\.\、:：)]|步骤\s*\d+[:：]?|预期\s*\d+[:：]?|[-*•])\s*")
EMPTY_FILL_PATTERN = re.compile(r"<fill\s*/>")


@dataclass
class CaseImportResult:
    sheet_name: str
    row_number: int
    title: str
    module_name: str
    payload: dict[str, Any]
    request: dict[str, Any]
    status: str
    testcase_id: int | None = None
    response: dict[str, Any] | None = None
    error: str | None = None


@dataclass
class SheetImportResult:
    sheet_name: str
    module_name: str
    module_id: int
    parsed_count: int = 0
    skipped_count: int = 0
    success_count: int = 0
    failed_count: int = 0
    cases: list[CaseImportResult] = field(default_factory=list)


@dataclass
class WorkbookImportResult:
    workbook_path: str
    product_id: int
    dry_run: bool
    source: dict[str, Any]
    generated_at: str
    report_path: str
    total_sheets: int
    parsed_count: int
    skipped_count: int
    success_count: int
    failed_count: int
    sheets: list[SheetImportResult]
    verification: dict[str, Any] | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "workbook_path": self.workbook_path,
            "product_id": self.product_id,
            "dry_run": self.dry_run,
            "source": self.source,
            "generated_at": self.generated_at,
            "report_path": self.report_path,
            "total_sheets": self.total_sheets,
            "parsed_count": self.parsed_count,
            "skipped_count": self.skipped_count,
            "success_count": self.success_count,
            "failed_count": self.failed_count,
            "sheets": [asdict(sheet) for sheet in self.sheets],
            "verification": self.verification,
        }


def project_root() -> Path:
    return find_project_root(Path(__file__))


class ZentaoTestcaseClient:
    def __init__(
        self,
        *,
        base_url: str,
        account: str | None = None,
        password: str | None = None,
        token: str | None = None,
        product_id: int = DEFAULT_PRODUCT_ID,
        timeout: int = 30,
    ):
        load_dotenv_file()
        self.base_url = base_url.rstrip("/")
        self.account = account
        self.password = password
        self.token = token
        self.product_id = product_id or DEFAULT_PRODUCT_ID
        self.timeout = timeout
        self.session = requests.Session()
        self.session.headers.update({"Accept": "application/json"})

    def ensure_token(self) -> str:
        if self.token:
            return self.token
        if not self.account or not self.password:
            raise RuntimeError("未配置 ZenTao 登录账号或密码，也没有可用 token。")
        response = self.session.post(
            f"{self.base_url}/users/login",
            json={"account": self.account, "password": self.password},
            timeout=self.timeout,
        )
        response.raise_for_status()
        payload = response.json()
        if str(payload.get("status", "")).lower() != "success" or not payload.get("token"):
            raise RuntimeError(f"ZenTao 登录失败，返回内容：{payload}")
        self.token = str(payload["token"])
        return self.token

    def create_testcase(self, payload: dict[str, Any]) -> dict[str, Any]:
        token = self.ensure_token()
        request_meta = {
            "method": "POST",
            "url": f"{self.base_url}/testcases",
            "headers": {"token": _mask_token(token)},
            "json": payload,
        }
        response = self.session.post(
            f"{self.base_url}/testcases",
            json=payload,
            headers={"token": token},
            timeout=self.timeout,
        )
        if response.status_code in {401, 403}:
            raise RuntimeError("ZenTao 创建测试用例被拒绝，请确认 token 请求头和账号权限是否正确。")
        response.raise_for_status()
        data = response.json()
        if str(data.get("status", "")).lower() != "success" or not data.get("id"):
            raise RuntimeError(f"ZenTao 创建测试用例失败，返回内容：{data}")
        return {
            "request": request_meta,
            "response": {
                "status_code": response.status_code,
                "body": data,
            },
            "testcase_id": int(data["id"]),
        }

    def list_product_testcases(self, *, rec_per_page: int = 1000, page_id: int = 1) -> dict[str, Any]:
        token = self.ensure_token()
        url = f"{self.base_url}/products/{self.product_id}/testcases"
        response = self.session.get(
            url,
            params={"recPerPage": rec_per_page, "pageID": page_id},
            headers={"token": token},
            timeout=self.timeout,
        )
        if response.status_code in {401, 403}:
            raise RuntimeError("ZenTao 查询测试用例列表被拒绝，请确认 token 请求头和账号权限是否正确。")
        response.raise_for_status()
        data = response.json()
        if str(data.get("status", "")).lower() != "success":
            raise RuntimeError(f"ZenTao 查询测试用例列表失败，返回内容：{data}")
        return {
            "request": {
                "method": "GET",
                "url": url,
                "headers": {"token": _mask_token(token)},
                "params": {"recPerPage": rec_per_page, "pageID": page_id},
            },
            "response": {
                "status_code": response.status_code,
                "body": data,
            },
        }


def import_testcases_from_excel(
    *,
    workbook_path: str | Path,
    base_url: str,
    account: str | None = None,
    password: str | None = None,
    token: str | None = None,
    product_id: int = DEFAULT_PRODUCT_ID,
    timeout: int = 30,
    dry_run: bool = False,
    output_path: str | Path | None = None,
    sheet_names: list[str] | None = None,
) -> WorkbookImportResult:
    workbook_file = Path(workbook_path).expanduser().resolve()
    if not workbook_file.exists():
        raise FileNotFoundError(f"未找到测试用例文件：{workbook_file}")
    if not dry_run and not base_url.strip():
        raise RuntimeError("未配置 ZenTao BASE_URL，无法执行真实测试用例上传。")

    workbook = _load_workbook_with_style_repair(workbook_file)
    selected_sheets = [name for name in workbook.sheetnames if not sheet_names or name in sheet_names]
    client = ZentaoTestcaseClient(
        base_url=base_url,
        account=account,
        password=password,
        token=token,
        product_id=product_id,
        timeout=timeout,
    )

    sheet_results: list[SheetImportResult] = []
    parsed_total = 0
    skipped_total = 0
    success_total = 0
    failed_total = 0

    for sheet_name in selected_sheets:
        worksheet = workbook[sheet_name]
        if worksheet.sheet_state != "visible":
            continue

        row_cells = list(worksheet.iter_rows())
        rows = [tuple(cell.value for cell in row) for row in row_cells]
        header_index, header_map = _detect_header_row(rows)
        if header_index is None or "title" not in header_map:
            continue

        module_name = _sheet_module_name(sheet_name, header_map, rows, header_index)
        module_id = _module_id(module_name)
        if module_id is None:
            raise RuntimeError(f"Sheet `{sheet_name}` 的模块 `{module_name}` 不在模块映射表里。")

        sheet_result = SheetImportResult(sheet_name=sheet_name, module_name=module_name, module_id=module_id)
        for offset, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
            row_cell_group = row_cells[offset - 1]
            if _row_is_empty(row):
                continue
            title = _cell_text(_pick(row, header_map, "title"))
            if not title:
                continue
            if _should_skip_row(row_cell_group, header_map):
                sheet_result.skipped_count += 1
                skipped_total += 1
                sheet_result.cases.append(
                    CaseImportResult(
                        sheet_name=sheet_name,
                        row_number=offset,
                        title=title,
                        module_name=module_name,
                        payload={},
                        request={},
                        status="skipped",
                        response={"status_code": None, "body": {"status": "skipped"}},
                        error="检测到删除线或下划线样式，按废弃用例跳过上传。",
                    )
                )
                continue

            payload = _build_testcase_payload(
                row=row,
                header_map=header_map,
                sheet_name=sheet_name,
                module_name=module_name,
                module_id=module_id,
                product_id=product_id,
            )
            sheet_result.parsed_count += 1
            parsed_total += 1

            case_result = CaseImportResult(
                sheet_name=sheet_name,
                row_number=offset,
                title=title,
                module_name=module_name,
                payload=payload,
                request={
                    "method": "POST",
                    "url": f"{client.base_url}/testcases",
                    "headers": {"token": _mask_token(client.token) if client.token else "<login-before-request>"},
                    "json": payload,
                },
                status="pending",
            )
            try:
                if dry_run:
                    case_result.status = "dry_run"
                    case_result.response = {
                        "status_code": None,
                        "body": {"status": "dry_run"},
                    }
                    sheet_result.success_count += 1
                    success_total += 1
                else:
                    upload_result = client.create_testcase(payload)
                    case_result.status = "success"
                    case_result.testcase_id = int(upload_result["testcase_id"])
                    case_result.request = upload_result["request"]
                    case_result.response = upload_result["response"]
                    sheet_result.success_count += 1
                    success_total += 1
            except Exception as exc:
                case_result.status = "failed"
                case_result.error = str(exc)
                if case_result.response is None:
                    case_result.response = {
                        "status_code": None,
                        "body": {"status": "fail", "message": str(exc)},
                    }
                sheet_result.failed_count += 1
                failed_total += 1
            sheet_result.cases.append(case_result)

        if sheet_result.parsed_count:
            sheet_results.append(sheet_result)

    report_file = _report_output_path(workbook_file, output_path)
    result = WorkbookImportResult(
        workbook_path=str(workbook_file),
        product_id=product_id,
        dry_run=dry_run,
        source={"type": "api", "base_url": base_url, "product_id": product_id},
        generated_at=datetime.now().isoformat(timespec="seconds"),
        report_path=str(report_file),
        total_sheets=len(sheet_results),
        parsed_count=parsed_total,
        skipped_count=skipped_total,
        success_count=success_total,
        failed_count=failed_total,
        sheets=sheet_results,
    )
    if not dry_run:
        result.verification = _verify_created_testcases(result, client)
    report_file.write_text(json.dumps(result.to_dict(), ensure_ascii=False, indent=2), encoding="utf-8")
    return result


def _report_output_path(workbook_path: Path, output_path: str | Path | None) -> Path:
    if output_path:
        output = Path(output_path)
    else:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output = project_root() / "artifacts" / "testcase_imports" / f"{workbook_path.stem}_zentao_import_{timestamp}.json"
    output.parent.mkdir(parents=True, exist_ok=True)
    if not output.exists():
        return output
    stem = output.stem
    suffix = output.suffix
    for index in range(2, 10_000):
        candidate = output.with_name(f"{stem}_{index}{suffix}")
        if not candidate.exists():
            return candidate
    raise RuntimeError(f"无法生成不重名的导入结果文件：{output}")


def _detect_header_row(rows: list[tuple[Any, ...]]) -> tuple[int | None, dict[str, int]]:
    best_index: int | None = None
    best_map: dict[str, int] = {}
    best_score = -1
    for index, row in enumerate(rows[:20]):
        normalized = [_normalize_header(cell) for cell in row]
        header_map: dict[str, int] = {}
        for field, aliases in HEADER_ALIASES.items():
            for cell_index, value in enumerate(normalized):
                if value in aliases:
                    header_map[field] = cell_index
                    break
        score = len(header_map)
        if score > best_score:
            best_index = index
            best_map = header_map
            best_score = score
    if best_score <= 0:
        return None, {}
    return best_index, best_map


def _sheet_module_name(
    sheet_name: str,
    header_map: dict[str, int],
    rows: list[tuple[Any, ...]],
    header_index: int,
) -> str:
    if "module" in header_map:
        for row in rows[header_index + 1 :]:
            candidate = _cell_text(_pick(row, header_map, "module"))
            if candidate:
                return candidate
    return sheet_name.strip()


def _module_id(module_name: str) -> int | None:
    compact = _normalize_module_name(module_name)
    if compact in MODULE_ID_MAP:
        return MODULE_ID_MAP[compact]

    for canonical_name, aliases in MODULE_NAME_ALIASES.items():
        normalized_aliases = {_normalize_module_name(alias) for alias in aliases | {canonical_name}}
        if compact in normalized_aliases:
            return MODULE_ID_MAP.get(canonical_name)
        if any(alias and alias in compact for alias in normalized_aliases):
            return MODULE_ID_MAP.get(canonical_name)
        if any(compact and compact in alias for alias in normalized_aliases):
            return MODULE_ID_MAP.get(canonical_name)
    return None


def _normalize_module_name(value: str) -> str:
    return value.replace(" ", "").replace("-", "").replace("_", "").strip()


def _build_testcase_payload(
    *,
    row: tuple[Any, ...],
    header_map: dict[str, int],
    sheet_name: str,
    module_name: str,
    module_id: int,
    product_id: int,
) -> dict[str, Any]:
    steps = _split_lines(_cell_text(_pick(row, header_map, "steps")))
    expects = _split_lines(_cell_text(_pick(row, header_map, "expects")))
    steps, expects = _align_steps_and_expects(steps, expects)
    payload: dict[str, Any] = {
        "product": product_id,
        "productID": product_id,
        "title": _cell_text(_pick(row, header_map, "title")),
        "module": module_id,
        "pri": _parse_int(_pick(row, header_map, "pri")) or 3,
        "type": _normalize_case_type(_cell_text(_pick(row, header_map, "type"))),
        "precondition": _cell_text(_pick(row, header_map, "precondition")),
        "steps": steps,
        "expects": expects,
        "stepType": ["step"] * len(steps),
    }

    for field in ("story", "project", "execution"):
        value = _parse_int(_pick(row, header_map, field))
        if value is not None:
            payload[field] = value

    if not payload["precondition"]:
        payload.pop("precondition")
    if not steps:
        payload.pop("steps")
        payload.pop("expects")
        payload.pop("stepType")
    if not payload["type"]:
        payload["type"] = "feature"
    if not payload["title"]:
        raise RuntimeError(f"Sheet `{sheet_name}` 模块 `{module_name}` 存在空标题用例。")
    return payload


def _normalize_header(value: Any) -> str:
    text = _cell_text(value).lower()
    return text.replace(" ", "")


def _normalize_case_type(value: str) -> str:
    compact = value.strip().lower()
    if not compact:
        return "feature"
    return CASE_TYPE_MAP.get(compact, CASE_TYPE_MAP.get(value.strip(), "feature"))


def _pick(row: tuple[Any, ...], header_map: dict[str, int], field: str) -> Any:
    index = header_map.get(field)
    if index is None or index >= len(row):
        return None
    return row[index]


def _row_is_empty(row: tuple[Any, ...]) -> bool:
    return not any(_cell_text(cell) for cell in row)


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_int(value: Any) -> int | None:
    text = _cell_text(value)
    if not text:
        return None
    match = re.search(r"\d+", text)
    if match and not text.replace(".", "", 1).isdigit():
        return int(match.group(0))
    try:
        return int(float(text))
    except ValueError:
        return None


def _split_lines(value: str) -> list[str]:
    if not value:
        return []
    result = []
    for part in STEP_SPLIT_PATTERN.split(value):
        cleaned = LINE_PREFIX_PATTERN.sub("", part).strip()
        if cleaned:
            result.append(cleaned)
    return result


def _align_steps_and_expects(steps: list[str], expects: list[str]) -> tuple[list[str], list[str]]:
    if not steps and expects:
        steps = [f"步骤{i}" for i in range(1, len(expects) + 1)]
    if steps and not expects:
        expects = [""] * len(steps)
    if len(steps) == len(expects):
        return steps, expects
    if steps and expects and len(expects) < len(steps):
        padded_expects = [""] * (len(steps) - len(expects)) + expects
        return steps, padded_expects
    max_len = max(len(steps), len(expects))
    padded_steps = steps + [""] * (max_len - len(steps))
    padded_expects = expects + [""] * (max_len - len(expects))
    return padded_steps, padded_expects


def _mask_token(token: str | None) -> str:
    if not token:
        return ""
    if len(token) <= 8:
        return "*" * len(token)
    return f"{token[:4]}{'*' * (len(token) - 8)}{token[-4:]}"


def _should_skip_row(row_cells: tuple[Any, ...], header_map: dict[str, int]) -> bool:
    candidate_fields = ("module", "title", "precondition", "steps", "expects")
    checked_indexes = []
    for field in candidate_fields:
        index = header_map.get(field)
        if index is not None and index < len(row_cells):
            checked_indexes.append(index)
    if not checked_indexes:
        checked_indexes = list(range(min(len(row_cells), 6)))
    for index in checked_indexes:
        cell = row_cells[index]
        font = getattr(cell, "font", None)
        if not font:
            continue
        if bool(font.strike):
            return True
        underline = getattr(font, "underline", None)
        if underline not in (None, "", "none"):
            return True
    return False


def _verify_created_testcases(result: WorkbookImportResult, client: ZentaoTestcaseClient) -> dict[str, Any]:
    created_cases = [
        case
        for sheet in result.sheets
        for case in sheet.cases
        if case.status == "success" and case.testcase_id is not None
    ]
    created_ids = [int(case.testcase_id) for case in created_cases]
    created_titles = [case.title for case in created_cases]
    try:
        listing = client.list_product_testcases()
        response_body = listing["response"]["body"]
        listed_cases = response_body.get("testcases") or []
        listed_ids = {
            _parse_int(item.get("id"))
            for item in listed_cases
            if isinstance(item, dict) and _parse_int(item.get("id")) is not None
        }
        listed_titles = {
            str(item.get("title", "")).strip()
            for item in listed_cases
            if isinstance(item, dict) and str(item.get("title", "")).strip()
        }
        missing_ids = [case_id for case_id in created_ids if case_id not in listed_ids]
        missing_titles = [title for title in created_titles if title not in listed_titles]
        matched_ids = [case_id for case_id in created_ids if case_id in listed_ids]
        matched_titles = [title for title in created_titles if title in listed_titles]
        return {
            "verified": len(missing_ids) == 0,
            "created_count": len(created_ids),
            "product_list_total": int((response_body.get("pager") or {}).get("recTotal") or 0),
            "matched_by_id_count": len(matched_ids),
            "matched_by_title_count": len(matched_titles),
            "missing_ids": missing_ids,
            "missing_titles": missing_titles,
            "request": listing["request"],
            "response": listing["response"],
        }
    except Exception as exc:
        return {
            "verified": False,
            "created_count": len(created_ids),
            "error": str(exc),
        }


def _load_workbook_with_style_repair(workbook_file: Path):
    try:
        return load_workbook(workbook_file, data_only=True)
    except TypeError as exc:
        message = str(exc)
        if "openpyxl.styles.fills.Fill" not in message:
            raise
        repaired_file = _repair_workbook_styles(workbook_file)
        return load_workbook(repaired_file, data_only=True)


def _repair_workbook_styles(workbook_file: Path) -> Path:
    temp_dir = Path(tempfile.mkdtemp(prefix="qatoolkit_xlsx_fix_"))
    repaired_file = temp_dir / workbook_file.name
    shutil.copy2(workbook_file, repaired_file)
    with ZipFile(repaired_file, "a") as archive:
        styles_xml = archive.read("xl/styles.xml").decode("utf-8", errors="replace")
        sanitized = EMPTY_FILL_PATTERN.sub(
            "<fill><patternFill patternType=\"none\"/></fill>",
            styles_xml,
        )
        archive.writestr("xl/styles.xml", sanitized)
    return repaired_file
